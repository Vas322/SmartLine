"""Tests for the Excel export service."""
from datetime import timedelta
from decimal import Decimal

from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook

from core.models import Activity, Player, TelegramMessage
from reports.services.excel_exporter import export_activities_excel


class ExportActivitiesExcelTests(TestCase):
    def setUp(self):
        self.player = Player.objects.create(nickname="Swettka")
        self.message = TelegramMessage.objects.create(
            telegram_chat_id=10,
            telegram_message_id=20,
            text="+1 | деф | Swettka | Первая волна",
            message_date=timezone.now(),
            status=TelegramMessage.Status.PROCESSED,
        )
        self.activity = Activity.objects.create(
            player=self.player,
            telegram_message=self.message,
            amount=Decimal("1.5"),
            activity_type=Activity.ActivityType.DEF,
            description="Первая волна",
        )

    def _read_rows(self, date_from, date_to):
        stream = export_activities_excel(date_from, date_to)
        workbook = load_workbook(stream, read_only=True)
        sheet = workbook.active
        return list(sheet.iter_rows(values_only=True))

    def test_export_contains_expected_columns_and_values(self):
        date_from = timezone.now() - timedelta(days=1)
        date_to = timezone.now() + timedelta(days=1)
        rows = self._read_rows(date_from, date_to)

        headers = rows[0]
        self.assertEqual(
            headers,
            (
                "Дата",
                "Игрок",
                "Тип",
                "Количество часов",
                "Минуты",
                "Оплачиваемые часы",
                "Описание",
                "Исходное сообщение",
                "Telegram message id",
            ),
        )

        self.assertEqual(len(rows), 2)
        row = rows[1]
        self.assertEqual(row[1], "Swettka")
        self.assertEqual(row[2], "DEF")
        self.assertEqual(row[3], "1.5")
        self.assertEqual(row[4], 90)
        self.assertEqual(row[5], "1.5")
        self.assertEqual(row[6], "Первая волна")
        self.assertEqual(row[7], "+1 | деф | Swettka | Первая волна")
        self.assertEqual(row[8], 20)

    def test_farm_paid_hours_are_zero(self):
        self.activity.activity_type = Activity.ActivityType.FARM
        self.activity.save(update_fields=["activity_type"])

        date_from = timezone.now() - timedelta(days=1)
        date_to = timezone.now() + timedelta(days=1)
        rows = self._read_rows(date_from, date_to)
        self.assertEqual(rows[1][5], "0")

    def test_cast_paid_hours_are_counted(self):
        self.activity.activity_type = Activity.ActivityType.CAST
        self.activity.has_cast = True
        self.activity.save(update_fields=["activity_type", "has_cast"])

        date_from = timezone.now() - timedelta(days=1)
        date_to = timezone.now() + timedelta(days=1)
        rows = self._read_rows(date_from, date_to)
        self.assertEqual(rows[1][2], "CAST")
        self.assertEqual(rows[1][5], "1.5")

    def test_farm_plus_cast_paid_hours_are_counted(self):
        self.activity.activity_type = Activity.ActivityType.FARM
        self.activity.has_cast = True
        self.activity.save(update_fields=["activity_type", "has_cast"])

        date_from = timezone.now() - timedelta(days=1)
        date_to = timezone.now() + timedelta(days=1)
        rows = self._read_rows(date_from, date_to)
        self.assertEqual(rows[1][2], "FARM+CAST")
        self.assertEqual(rows[1][5], "1.5")

    def test_def_plus_cast_export_type(self):
        self.activity.activity_type = Activity.ActivityType.DEF
        self.activity.has_cast = True
        self.activity.save(update_fields=["activity_type", "has_cast"])

        date_from = timezone.now() - timedelta(days=1)
        date_to = timezone.now() + timedelta(days=1)
        rows = self._read_rows(date_from, date_to)
        self.assertEqual(rows[1][2], "DEF+CAST")

    def test_export_respects_period(self):
        date_from = timezone.now() - timedelta(days=30)
        date_to = timezone.now() - timedelta(days=29)
        rows = self._read_rows(date_from, date_to)
        self.assertEqual(len(rows), 1)  # only the header row

    def test_multiple_activities_on_one_message_are_exported(self):
        player2 = Player.objects.create(nickname="Pocomaxa")
        Activity.objects.create(
            player=player2,
            telegram_message=self.message,
            amount=Decimal("1.5"),
            activity_type=Activity.ActivityType.DEF,
            description="Первая волна",
        )

        date_from = timezone.now() - timedelta(days=1)
        date_to = timezone.now() + timedelta(days=1)
        rows = self._read_rows(date_from, date_to)

        self.assertEqual(len(rows), 3)  # header + two activities
        nicknames = {row[1] for row in rows[1:]}
        self.assertEqual(nicknames, {"Swettka", "Pocomaxa"})
        # source text stays reachable through the FK on both rows
        for row in rows[1:]:
            self.assertEqual(row[7], "+1 | деф | Swettka | Первая волна")
            self.assertEqual(row[8], 20)
